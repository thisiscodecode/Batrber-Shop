from datetime import date, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from api.models.booking import Booking
from api.models.user import User
from api.models.service import Service


async def generate_bookings_excel(
    db: AsyncSession,
    date_from: date | None = None,
    date_to: date | None = None,
    status: str | None = None,
) -> bytes:
    query = (
        select(Booking)
        .options(selectinload(Booking.user), selectinload(Booking.service))
        .order_by(Booking.created_at.desc())
    )

    if date_from:
        query = query.where(Booking.date >= date_from)
    if date_to:
        query = query.where(Booking.date <= date_to)
    if status:
        query = query.where(Booking.status == status)

    result = await db.execute(query)
    bookings = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["ID", "Client Name", "Phone", "Telegram ID", "Service", "Date", "Time", "Status", "Note", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, booking in enumerate(bookings, 2):
        values = [
            booking.id,
            booking.user.name if booking.user else "",
            booking.user.phone if booking.user else "",
            booking.user.telegram_id if booking.user else "",
            booking.service.title if booking.service else "",
            booking.date.isoformat() if booking.date else "",
            booking.time.strftime("%H:%M") if booking.time else "",
            booking.status,
            booking.note or "",
            booking.created_at.strftime("%Y-%m-%d %H:%M") if booking.created_at else "",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
